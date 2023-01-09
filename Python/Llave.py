import openpyxl
import time
import re
from openpyxl.styles import Color, colors

print("¿Cuál es la ubicación que buscas?")
ubicacion = input()

ubicacion_frag = ubicacion.split(" ")

inicio = time.time()

def obtener_ubicacion(ubi_frg):
    #numero_piso = re.split(r'[A-Z]', ubi_frg[2])
    piso = ubi_frg[1][1:2]
    
    return f"E:\\Carpetas escritorio\\desarrollo\\IT\\Llaves\\{ubi_frg[0]}\\Piso {piso}.xlsx"

path = obtener_ubicacion(ubicacion_frag)

wb_obj = openpyxl.load_workbook(path)

isla = ubicacion_frag[2][:1]
puesto = ubicacion_frag[2][1:]

sheets = wb_obj.get_sheet_names()
islas = sheets[0].split(" ")[2]

if (re.compile('[A-M]').match(isla)):
    ala = "Norte"
else: 
    ala = "Sur"

if (ala == "Norte"):
    sheet = wb_obj.get_sheet_by_name(sheets[0])
else:
    sheet = wb_obj.get_sheet_by_name(sheets[1])

max_col = sheet.max_column
max_row = sheet.max_row

isla_puesto = isla + "-" + puesto

for i in range(1, max_col + 1):
    for j in range(1, max_row + 1):
        cell_obj = sheet.cell(row = j, column = i)
        value = cell_obj.value
        if (value == isla_puesto):
            print(ala, value)
            cell_obj.fill = "#FFFFFF"
            print(cell_obj.fill)
            break

fin = time.time()
print(fin-inicio)